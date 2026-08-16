"""
Import von Aurora-Marktdaten (engine/io_aurora.py).

Der Import baut aus den Aurora-Dateien ein Marktpreisszenario. Die harte
Bedingung: MONATSDATEN. Zweiseitiger CfD und Abschoepfung sind
abgeschnittene Funktionen des Marktwerts - ob eine Rueckzahlung
entsteht, entscheidet sich in einzelnen Monaten und ist aus einem
Jahresmittel nicht rekonstruierbar. Ein Import ohne Monatsaufloesung
wuerde diese Rechnung stillschweigend entwerten und bricht deshalb ab.

Die Testdateien bilden die Spaltenstruktur der Aurora-Exporte nach; die
Zahlen sind so gewaehlt, dass jede Erwartung von Hand nachrechenbar ist.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from engine.io_aurora import (  # noqa: E402
    AuroraImportFehler,
    importiere_aurora,
    technologien,
    vorschlag_solar,
)

#: Erzeugung je Monat (TWh) - sommerlastig wie eine PV-Anlage.
ERZEUGUNG = [0.3, 0.5, 1.0, 1.3, 1.5, 1.6, 1.6, 1.4, 1.1, 0.8, 0.4, 0.3]
#: Capture Price je Monat (EUR/MWh) - im Sommer niedrig.
PREIS = [70.0, 65.0, 55.0, 45.0, 35.0, 30.0, 32.0, 38.0, 50.0, 60.0, 68.0, 72.0]


def _tech_monat_csv(
    jahre=(2030, 2031), monate=range(1, 13), mit_erzeugung: bool = True,
    gruppen=(("Solar", "Solar PV"), ("Wind", "Onshore wind")),
) -> tuple[bytes, str]:
    zeilen = []
    for jahr in jahre:
        for gruppe, untergruppe in gruppen:
            faktor = 1.0 if gruppe == "Solar" else 1.2
            for monat in monate:
                zeilen.append(
                    {
                        "Year": jahr,
                        "Month": monat,
                        "Group": gruppe,
                        "Subgroup": untergruppe,
                        "Capacity": 10.0,
                        "New build capacity": 0.5,
                        "Capture price curtailed below zero": (
                            PREIS[monat - 1] * faktor * 0.9
                        ),
                        "Generation": ERZEUGUNG[monat - 1] if mit_erzeugung else None,
                        "Uncurtailed capture price": PREIS[monat - 1] * faktor,
                        "Curtailment rate below zero - 15 minute rule": 0.20,
                        "Curtailment rate below zero - 1 hour rule with average "
                        "hourly price below zero": 0.15,
                        "Curtailment rate below zero - 6 hour rule with average "
                        "hourly price below zero": 0.10,
                        "Total CAPEX cost": 600.0,
                        "Total OPEX cost": 12.0,
                        "Imbalance Cost": 2.0,
                    }
                )
    df = pd.DataFrame(zeilen)
    if not mit_erzeugung:
        df = df.drop(columns=["Generation"])
    return df.to_csv(index=False).encode("utf-8"), "technologie_monat.csv"


def _tech_jahr_csv(jahre=(2030, 2031), preis: float | None = None) -> tuple[bytes, str]:
    """Jahresdatei - ohne Monatsspalte, wie der echte Export."""
    gewichtet = sum(p * e for p, e in zip(PREIS, ERZEUGUNG, strict=True)) / sum(
        ERZEUGUNG
    )
    zeilen = [
        {
            "Year": jahr,
            "Group": "Solar",
            "Subgroup": "Solar PV",
            "Capacity": 10.0,
            "Generation": sum(ERZEUGUNG),
            "Uncurtailed capture price": preis if preis is not None else gewichtet,
            "Capture price curtailed below zero": (
                (preis if preis is not None else gewichtet) * 0.9
            ),
            "Curtailment rate below zero - 1 hour rule with average hourly price "
            "below zero": 0.15,
            "Curtailment rate below zero - 6 hour rule with average hourly price "
            "below zero": 0.10,
        }
        for jahr in jahre
    ]
    return (
        pd.DataFrame(zeilen).to_csv(index=False).encode("utf-8"),
        "technologie_jahr.csv",
    )


def _system_jahr_csv(jahre=(2025, 2030), rate: float = 0.02) -> tuple[bytes, str]:
    zeilen = [
        {
            "year": jahr,
            "Baseload price": 60.0,
            "Gas price": 25.0,
            "Carbon price": 90.0,
            "Total demand": 70.0,
            "EUR/GBP": 0.85,
            "USD/EUR": 1.08,
            "EUR Inflation (Calendar Year)": (1 + rate) ** (jahr - 2025),
        }
        for jahr in range(jahre[0], jahre[1] + 1)
    ]
    return pd.DataFrame(zeilen).to_csv(index=False).encode("utf-8"), "system_jahr.csv"


class TestMonatsdatenSindPflicht:
    """Die Bedingung, an der der Import scheitern SOLL."""

    def test_jahresdatei_im_monatsfeld_wird_abgewiesen(self):
        """Wer die Jahresdatei laedt, bekommt keinen halben Import,
        sondern eine Fehlermeldung - sonst faende sich der Fehler erst
        in der Rendite wieder."""
        with pytest.raises(AuroraImportFehler, match="Monat"):
            importiere_aurora(
                name="Aurora Test", technologie_monat=_tech_jahr_csv()
            )

    def test_unvollstaendige_monatsreihe_zaehlt_nicht(self):
        """Zehn Monate sind keine Monatsreihe: Die zwoelf Werte stuenden
        sonst verschoben in der Kurve."""
        with pytest.raises(AuroraImportFehler, match="zwölf Monate"):
            importiere_aurora(
                name="Aurora Test",
                technologie_monat=_tech_monat_csv(monate=range(1, 11)),
            )

    def test_ein_vollstaendiges_jahr_genuegt(self):
        """Ein unvollstaendiges Jahr blockiert den Import nicht - es
        wird uebergangen, und der Hinweis nennt es."""
        inhalt, name = _tech_monat_csv(jahre=(2030,))
        teil = pd.read_csv(pd.io.common.BytesIO(inhalt))
        voll = _tech_monat_csv(jahre=(2031,))
        gemischt = pd.concat(
            [teil[teil["Month"] <= 8], pd.read_csv(pd.io.common.BytesIO(voll[0]))]
        )
        ergebnis = importiere_aurora(
            name="Aurora Test",
            technologie_monat=(
                gemischt.to_csv(index=False).encode("utf-8"), name
            ),
        )
        assert ergebnis.monatsjahre == 1
        assert 2031 in ergebnis.szenario.marktwert_solar_ct_kwh_je_monat
        assert 2030 not in ergebnis.szenario.marktwert_solar_ct_kwh_je_monat
        assert any("Unvollständige" in h for h in ergebnis.hinweise)

    def test_name_ist_pflicht(self):
        with pytest.raises(AuroraImportFehler, match="Namen"):
            importiere_aurora(name="  ", technologie_monat=_tech_monat_csv())


class TestKurvenAusDenDaten:
    def test_capture_price_wird_zu_ct_je_kwh(self):
        """Aurora rechnet in EUR/MWh, das Modell in ct/kWh."""
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv()
        )
        monate = ergebnis.szenario.marktwert_solar_ct_kwh_je_monat[2030]
        assert monate == pytest.approx([p / 10 for p in PREIS])

    def test_jahreswert_ist_das_erzeugungsgewichtete_mittel(self):
        """Nicht der einfache Durchschnitt: Fuer PV waere er zu hoch,
        weil die ertragsstarken Monate die preisschwachen sind - und die
        Engine bildet beim Verdichten dieselbe Groesse."""
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv()
        )
        erwartet = sum(p * e for p, e in zip(PREIS, ERZEUGUNG, strict=True)) / sum(
            ERZEUGUNG
        )
        jahreswert = ergebnis.szenario.marktwert_solar_ct_kwh_je_kalenderjahr[2030]
        assert jahreswert == pytest.approx(erwartet / 10)
        assert jahreswert < sum(PREIS) / 12 / 10

    def test_beide_abregelungsregeln_werden_gefuellt(self):
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv()
        )
        szenario = ergebnis.szenario
        assert szenario.erzeugungsmenge_negativ_6h_pct_je_monat[2030] == (
            pytest.approx([0.10] * 12)
        )
        assert szenario.erzeugungsmenge_negativ_1h_pct_je_monat[2030] == (
            pytest.approx([0.15] * 12)
        )
        # Die 1h-Regel erfasst mehr Menge als die 6h-Regel.
        assert (
            szenario.erzeugungsmenge_negativ_1h_pct_je_kalenderjahr[2030]
            > szenario.erzeugungsmenge_negativ_6h_pct_je_kalenderjahr[2030]
        )

    def test_einspeisekurve_kommt_aus_der_erzeugung(self):
        """Der Ertragsverlauf steckt in den Daten und muss nicht
        geschaetzt werden."""
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv()
        )
        assert ergebnis.einspeisekurve_pct_je_monat == pytest.approx(
            [e / sum(ERZEUGUNG) for e in ERZEUGUNG]
        )
        assert sum(ergebnis.einspeisekurve_pct_je_monat) == pytest.approx(1.0)

    def test_ohne_erzeugungsspalte_bleibt_die_kurve_leer(self):
        ergebnis = importiere_aurora(
            name="Aurora Test",
            technologie_monat=_tech_monat_csv(mit_erzeugung=False),
        )
        assert ergebnis.einspeisekurve_pct_je_monat == []
        assert any("Erzeugungsspalte" in h for h in ergebnis.hinweise)
        # Ohne Gewichte bleibt der einfache Mittelwert.
        assert ergebnis.szenario.marktwert_solar_ct_kwh_je_kalenderjahr[2030] == (
            pytest.approx(sum(PREIS) / 12 / 10)
        )

    def test_gekuerzter_preis_auf_wunsch(self):
        """Voreingestellt ist der ungekuerzte Preis - die Abregelung
        bringt das Modell selbst ein."""
        ungekuerzt = importiere_aurora(
            name="A", technologie_monat=_tech_monat_csv()
        )
        gekuerzt = importiere_aurora(
            name="A", technologie_monat=_tech_monat_csv(), uncurtailed=False
        )
        assert gekuerzt.szenario.marktwert_solar_ct_kwh_je_monat[2030] == (
            pytest.approx(
                [w * 0.9 for w in ungekuerzt.szenario.marktwert_solar_ct_kwh_je_monat[2030]]
            )
        )


class TestTechnologieauswahl:
    def test_auswahl_listet_gruppe_und_untergruppe(self):
        assert technologien(*_tech_monat_csv()) == [
            "Solar · Solar PV", "Wind · Onshore wind",
        ]

    def test_vorschlag_findet_solar(self):
        assert vorschlag_solar(["Wind · Onshore wind", "Solar · Solar PV"]) == (
            "Solar · Solar PV"
        )
        assert vorschlag_solar(["Wind · Onshore wind"]) is None

    def test_ohne_angabe_wird_solar_gewaehlt(self):
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv()
        )
        assert ergebnis.technologie == "Solar · Solar PV"

    def test_andere_technologie_liefert_andere_preise(self):
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv(),
            technologie="Wind · Onshore wind",
        )
        assert ergebnis.szenario.marktwert_solar_ct_kwh_je_monat[2030] == (
            pytest.approx([p * 1.2 / 10 for p in PREIS])
        )

    def test_unbekannte_technologie_meldet_sich(self):
        with pytest.raises(AuroraImportFehler, match="Technologie"):
            importiere_aurora(
                name="Aurora Test", technologie_monat=_tech_monat_csv(),
                technologie="Kernfusion",
            )


class TestSystemdateiUndGegenprobe:
    def test_inflation_kommt_aus_dem_index(self):
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv(),
            system_jahr=_system_jahr_csv(rate=0.025),
        )
        assert ergebnis.inflation_basisjahr == 2025
        assert ergebnis.inflation_pct_pa == pytest.approx(0.025, rel=1e-6)

    def test_ohne_systemdatei_bleibt_die_inflation_unberuehrt(self):
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv()
        )
        assert ergebnis.inflation_basisjahr is None
        assert any("Systemdatei" in h for h in ergebnis.hinweise)

    def test_gegenprobe_schweigt_bei_uebereinstimmung(self):
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv(),
            technologie_jahr=_tech_jahr_csv(),
        )
        assert not [h for h in ergebnis.hinweise if "Gegenprobe" in h]

    def test_gegenprobe_meldet_abweichung(self):
        """Ein deutlich anderer Jahreswert deutet auf eine falsche
        Zuordnung hin - falsche Technologie, falscher Capture Price."""
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv(),
            technologie_jahr=_tech_jahr_csv(preis=80.0),
        )
        assert any("Gegenprobe" in h for h in ergebnis.hinweise)


class TestRobustheit:
    def test_prozentangaben_bei_den_quoten(self):
        """Manche Exporte schreiben die Abregelungsquote in Prozent."""
        inhalt, name = _tech_monat_csv()
        df = pd.read_csv(pd.io.common.BytesIO(inhalt))
        spalte = [s for s in df.columns if "6 hour" in s][0]
        df[spalte] = df[spalte] * 100
        ergebnis = importiere_aurora(
            name="Aurora Test",
            technologie_monat=(df.to_csv(index=False).encode("utf-8"), name),
        )
        assert ergebnis.szenario.erzeugungsmenge_negativ_6h_pct_je_monat[2030] == (
            pytest.approx([0.10] * 12)
        )
        assert any("Prozentangabe" in h for h in ergebnis.hinweise)

    def test_monatsnamen_statt_zahlen(self):
        """Aurora-Exporte schreiben den Monat mal als Zahl, mal als
        Namen."""
        inhalt, name = _tech_monat_csv()
        df = pd.read_csv(pd.io.common.BytesIO(inhalt))
        namen = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        df["Month"] = [namen[m - 1] for m in df["Month"]]
        ergebnis = importiere_aurora(
            name="Aurora Test",
            technologie_monat=(df.to_csv(index=False).encode("utf-8"), name),
        )
        assert ergebnis.szenario.marktwert_solar_ct_kwh_je_monat[2030] == (
            pytest.approx([p / 10 for p in PREIS])
        )

    def test_excel_wird_gelesen(self):
        import io as _io

        inhalt, _ = _tech_monat_csv()
        df = pd.read_csv(pd.io.common.BytesIO(inhalt))
        puffer = _io.BytesIO()
        with pd.ExcelWriter(puffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
        ergebnis = importiere_aurora(
            name="Aurora Test",
            technologie_monat=(puffer.getvalue(), "technologie_monat.xlsx"),
        )
        assert ergebnis.jahre == (2030, 2031)

    def test_fehlende_spalte_nennt_die_datei_und_die_groesse(self):
        inhalt, name = _tech_monat_csv()
        df = pd.read_csv(pd.io.common.BytesIO(inhalt))
        df = df.drop(columns=[s for s in df.columns if "capture price" in s.lower()])
        with pytest.raises(AuroraImportFehler, match="capture price"):
            importiere_aurora(
                name="Aurora Test",
                technologie_monat=(df.to_csv(index=False).encode("utf-8"), name),
            )


class TestZusammenspielMitDerEngine:
    def test_importiertes_szenario_rechnet_im_monatsmodus(
        self, project, global_assumptions
    ):
        """Der Zweck der Uebung: Mit den importierten Monatskurven muss
        die Abschoepfung rechenbar sein."""
        from engine import run_valuation
        from engine.models import PraemienModell, Zeitaufloesung

        ergebnis = importiere_aurora(
            name="Aurora Test",
            technologie_monat=_tech_monat_csv(jahre=range(2025, 2061)),
        )
        global_assumptions.marktpreisszenarien = [ergebnis.szenario]
        global_assumptions.einspeisekurve_pct_je_monat = (
            ergebnis.einspeisekurve_pct_je_monat
        )
        global_assumptions.zeitaufloesung = Zeitaufloesung.MONAT
        global_assumptions.praemien_modell = PraemienModell.EAG_TOLERANZBAND
        project.marktpreisszenario = "Aurora Test"
        project.nennleistung_kwp = 6000.0  # ueber der 5-MW-Schwelle

        cashflow = run_valuation(project, global_assumptions).cashflow.data
        # Winterpreise (7,2 ct) liegen ueber dem Toleranzband des
        # Zuschlagswerts von 7 ct (7 x 1,4 = 9,8 ct)? Nein - deshalb ist
        # hier keine Rueckzahlung zu erwarten, wohl aber eine Praemie.
        assert cashflow["erloes_praemie_eur"].sum() > 0
        assert cashflow["rueckzahlung_eur"].sum() == 0.0

    def test_kaputte_systemdatei_bricht_den_import_nicht_ab(self):
        """Die Systemdatei ist Beiwerk - sie liefert die Inflation, nicht
        die Kurven. Ein Fehler in ihr darf den Import nicht kosten."""
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv(),
            system_jahr=(b"nur;muell\n1;2\n", "system.csv"),
        )
        assert ergebnis.monatsjahre == 2
        assert any("Systemdatei" in h for h in ergebnis.hinweise)

    def test_monatliche_systemdatei_liefert_die_inflation_auch(self):
        inhalt, name = _system_jahr_csv(rate=0.03)
        df = pd.read_csv(pd.io.common.BytesIO(inhalt))
        monatlich = pd.concat([df.assign(Month=m) for m in range(1, 13)])
        ergebnis = importiere_aurora(
            name="Aurora Test", technologie_monat=_tech_monat_csv(),
            system_monat=(monatlich.to_csv(index=False).encode("utf-8"), name),
        )
        assert ergebnis.inflation_pct_pa == pytest.approx(0.03, rel=1e-6)


# ---------------------------------------------------------------------------
# Aurora-Arbeitsmappe (Market Forecast Data)
# ---------------------------------------------------------------------------

def _mappe_bauen(
    szenarien=("Central", "Low", "High"),
    monatsregel: str = "curtailment-below-zero",
    mit_jahresabregelung: bool = True,
    zweisprachig: bool = False,
    kopf_versatz: int = 0,
) -> bytes:
    """Baut eine Arbeitsmappe im Aurora-Format nach.

    Die Parameter bilden die Unterschiede zwischen den Jahrgaengen ab:
    zweisprachige Zusatzspalten, verschobene Kopfzeilen, anders benannte
    Abregelungsreihen. Genau daran darf der Import nicht scheitern.
    """
    import io as _io

    import openpyxl

    jahre = [2030, 2031]
    monate = list(range(1, 13))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Monatsblatt --------------------------------------------------
    ws = wb.create_sheet("Monthly prices")
    ws.cell(row=1, column=3, value="Title")
    ws.cell(row=1, column=4, value="Aurora TEST Power & Renewables")
    ws.cell(row=2, column=3, value="Geography")
    ws.cell(row=2, column=4, value="Testland")
    ws.cell(row=3, column=3, value="Currency")
    ws.cell(row=3, column=4, value="€, real 2025 - calendar year average")

    meta = 6 if zweisprachig else 5
    erste = meta + 1
    kopf_jahr = 7 + kopf_versatz
    ws.cell(row=kopf_jahr, column=meta, value="Calendar year")
    ws.cell(row=kopf_jahr + 2, column=meta, value="Month")
    spalte = erste
    for jahr in jahre:
        for monat in monate:
            ws.cell(row=kopf_jahr, column=spalte, value=jahr)
            ws.cell(row=kopf_jahr + 2, column=spalte, value=monat)
            spalte += 1

    zeile = kopf_jahr + 4

    def block(titel: str, reihen: list[tuple[str, dict[str, list[float]]]]):
        nonlocal zeile
        ws.cell(row=zeile, column=1, value=titel)
        zeile += 1
        for label, je_szenario in reihen:
            for i, szenario in enumerate(szenarien):
                # Beschriftung nur in der ersten Zeile der Gruppe - genau
                # wie Aurora es schreibt.
                if i == 0:
                    ws.cell(row=zeile, column=2, value=label)
                    if zweisprachig:
                        ws.cell(row=zeile, column=3, value=label + " (de)")
                ws.cell(row=zeile, column=meta - 2, value=szenario)
                ws.cell(row=zeile, column=meta - 1, value="Testland")
                ws.cell(row=zeile, column=meta, value="€/MWh")
                spalte = erste
                for jahr in jahre:
                    for monat in monate:
                        ws.cell(row=zeile, column=spalte,
                                value=je_szenario[szenario][monat - 1])
                        spalte += 1
                zeile += 1
            zeile += 1

    def reihe(faktor: float) -> dict[str, list[float]]:
        return {
            s: [p * faktor * (1.0 if s == "Central" else (0.7 if s == "Low" else 1.4))
                for p in PREIS]
            for s in szenarien
        }

    block("Results - baseload prices", [("Baseload", reihe(1.6))])
    block("Results - fixed solar PV", [
        ("Fixed solar PV uncurtailed capture price", reihe(1.0)),
        ("Fixed solar PV capture price curtailed-below-zero", reihe(1.2)),
        (f"Fixed solar PV {monatsregel}",
         {s: [10.0] * 12 for s in szenarien}),
    ])
    # Fussnote wie im Original: Sie sagt, auf welche Regel sich die
    # monatliche Abregelungsquote bezieht.
    ws.cell(row=zeile, column=2,
            value="Notes: Curtailment below zero is equivalent to the "
                  "15 minute rule")
    zeile += 2
    block("Results - tracking solar PV", [
        ("Tracking solar PV uncurtailed capture price", reihe(1.15)),
        (f"Tracking solar PV {monatsregel}", {s: [9.0] * 12 for s in szenarien}),
    ])

    # --- Jahresblaetter je Szenario -----------------------------------
    for szenario in szenarien:
        js = wb.create_sheet(szenario)
        js.cell(row=1, column=3, value="Scenario")
        js.cell(row=1, column=4, value=szenario)
        js.cell(row=4, column=3, value="Variable")
        for i, jahr in enumerate(jahre):
            js.cell(row=4, column=7 + i, value=jahr)
        z = 6

        def jahreszeile(block_titel, technologie, werte, zeile_nr):
            if block_titel:
                js.cell(row=zeile_nr, column=2, value=block_titel)
            js.cell(row=zeile_nr, column=3, value=technologie)
            js.cell(row=zeile_nr, column=5, value="Testland / Germany")
            js.cell(row=zeile_nr, column=6, value="€/MWh")
            for i, wert in enumerate(werte):
                js.cell(row=zeile_nr, column=7 + i, value=wert)

        faktor = 1.0 if szenario == "Central" else (0.7 if szenario == "Low" else 1.4)
        jahreszeile("Wholesale price", "Baseload", [80.0 * faktor] * 2, z)
        z += 2
        jahreszeile("Uncurtailed capture price", "Fixed solar PV",
                    [50.0 * faktor] * 2, z)
        z += 1
        jahreszeile(None, "Tracking solar PV", [57.5 * faktor] * 2, z)
        z += 2
        if mit_jahresabregelung:
            for regel, wert in (("1 hour", 18.0), ("6 hour", 12.0),
                                ("15 minute", 20.0)):
                jahreszeile(f"Renewables curtailment - {regel} rule",
                            "Fixed solar PV", [wert] * 2, z)
                z += 1
                jahreszeile(None, "Tracking solar PV", [wert * 0.9] * 2, z)
                z += 2

    puffer = _io.BytesIO()
    wb.save(puffer)
    return puffer.getvalue()


class TestArbeitsmappe:
    """Der zweite Weg: eine Datei statt vier CSV-Exporten."""

    def test_kopf_szenarien_und_technologien_werden_erkannt(self):
        from engine.io_aurora import lies_arbeitsmappe

        mappe = lies_arbeitsmappe(_mappe_bauen(), "Aurora_Q3_26_TEST.xlsx")
        assert mappe.geografie == "Testland"
        assert mappe.preisbasisjahr == 2025
        assert mappe.quartal == "Q3/26"
        assert mappe.szenarien == ["Central", "Low", "High"]
        assert mappe.technologien == ["Pult", "Tracker"]

    def test_drei_szenarien_auf_einmal(self):
        """Aus einer Mappe entstehen drei Marktpreisszenarien - genau
        die Struktur, mit der sich der Preispfad als Sensitivitaet
        rechnen laesst."""
        from engine.io_aurora import importiere_arbeitsmappe, lies_arbeitsmappe

        mappe = lies_arbeitsmappe(_mappe_bauen(), "Aurora_Q3_26_TEST.xlsx")
        ergebnisse = importiere_arbeitsmappe(mappe, "Aurora Q3/26")
        assert [e.szenario.name for e in ergebnisse] == [
            "Aurora Q3/26 · Pult · Central",
            "Aurora Q3/26 · Pult · Low",
            "Aurora Q3/26 · Pult · High",
        ]
        # Low liegt unter Central, High darueber - die Szenarien wurden
        # nicht vertauscht.
        werte = [e.szenario.marktwert_solar_ct_kwh_je_monat[2030][0] for e in ergebnisse]
        assert werte[1] < werte[0] < werte[2]

    def test_pult_und_tracker_sind_verschiedene_kurven(self):
        """Der Tracker erloest mehr - er trifft die preisschwachen
        Mittagsstunden weniger stark."""
        from engine.io_aurora import importiere_arbeitsmappe, lies_arbeitsmappe

        mappe = lies_arbeitsmappe(_mappe_bauen(), "Aurora_Q3_26_TEST.xlsx")
        pult = importiere_arbeitsmappe(mappe, "A", "Pult", ["Central"])[0]
        tracker = importiere_arbeitsmappe(mappe, "A", "Tracker", ["Central"])[0]
        assert pult.szenario.marktwert_solar_ct_kwh_je_kalenderjahr[2030] == (
            pytest.approx(5.0)
        )
        assert tracker.szenario.marktwert_solar_ct_kwh_je_kalenderjahr[2030] == (
            pytest.approx(5.75)
        )
        assert "Pult" in pult.szenario.name and "Tracker" in tracker.szenario.name

    def test_grosshandelspreis_wird_uebernommen(self):
        from engine.io_aurora import importiere_arbeitsmappe, lies_arbeitsmappe

        mappe = lies_arbeitsmappe(_mappe_bauen(), "Aurora_Q3_26_TEST.xlsx")
        szenario = importiere_arbeitsmappe(mappe, "A", "Pult", ["Central"])[0].szenario
        assert szenario.baseload_ct_kwh_je_kalenderjahr[2030] == pytest.approx(8.0)
        assert szenario.baseload_ct_kwh_je_monat[2030] == pytest.approx(
            [p * 1.6 / 10 for p in PREIS]
        )
        # Der Baseload liegt ueber dem Marktwert Solar - der Abstand ist
        # der Kannibalisierungseffekt.
        assert (
            szenario.baseload_ct_kwh_je_kalenderjahr[2030]
            > szenario.marktwert_solar_ct_kwh_je_kalenderjahr[2030]
        )

    def test_abregelung_folgt_dem_monatsprofil_auf_jahresniveau(self):
        """Das Monatsblatt fuehrt nur EINE Abregelungsreihe (hier die
        15-Minuten-Regel), das Jahresblatt drei. Die 6h-Reihe entsteht
        aus dem Monatsprofil und dem Jahresniveau der 6h-Regel."""
        from engine.io_aurora import importiere_arbeitsmappe, lies_arbeitsmappe

        mappe = lies_arbeitsmappe(_mappe_bauen(), "Aurora_Q3_26_TEST.xlsx")
        szenario = importiere_arbeitsmappe(mappe, "A", "Pult", ["Central"])[0].szenario
        assert szenario.erzeugungsmenge_negativ_6h_pct_je_kalenderjahr[2030] == (
            pytest.approx(0.12)
        )
        assert szenario.erzeugungsmenge_negativ_1h_pct_je_kalenderjahr[2030] == (
            pytest.approx(0.18)
        )
        # Monatsreihe: 10 % skaliert auf das Niveau der jeweiligen Regel
        # (Referenz ist die 15-Minuten-Regel mit 20 %).
        assert szenario.erzeugungsmenge_negativ_6h_pct_je_monat[2030] == (
            pytest.approx([0.10 * 12.0 / 20.0] * 12)
        )
        assert szenario.erzeugungsmenge_negativ_1h_pct_je_monat[2030] == (
            pytest.approx([0.10 * 18.0 / 20.0] * 12)
        )

    def test_monatsreihe_mit_eigener_regelbezeichnung(self):
        """Aeltere Ausgaben nennen die Regel in der Beschriftung
        („curtailment % - 1 hour rule“) - dann gilt sie als Referenz."""
        from engine.io_aurora import importiere_arbeitsmappe, lies_arbeitsmappe

        mappe = lies_arbeitsmappe(
            _mappe_bauen(monatsregel="curtailment % - 1 hour rule"), "Aurora.xlsx"
        )
        szenario = importiere_arbeitsmappe(mappe, "A", "Pult", ["Central"])[0].szenario
        assert szenario.erzeugungsmenge_negativ_1h_pct_je_monat[2030] == (
            pytest.approx([0.10] * 12)
        )
        assert szenario.erzeugungsmenge_negativ_6h_pct_je_monat[2030] == (
            pytest.approx([0.10 * 12.0 / 18.0] * 12)
        )

    def test_zweisprachige_mappe_und_verschobene_kopfzeilen(self):
        """Aurora verschiebt zwischen den Ausgaben Kopfzeilen und fuegt
        deutsche Zweitspalten hinzu - beides darf nichts aendern."""
        from engine.io_aurora import importiere_arbeitsmappe, lies_arbeitsmappe

        mappe = lies_arbeitsmappe(
            _mappe_bauen(zweisprachig=True, kopf_versatz=-1), "Aurora.xlsx"
        )
        szenario = importiere_arbeitsmappe(mappe, "A", "Pult", ["Central"])[0].szenario
        assert szenario.marktwert_solar_ct_kwh_je_monat[2030] == pytest.approx(
            [p / 10 for p in PREIS]
        )

    def test_ohne_jahresabregelung_gilt_die_monatsquote(self):
        from engine.io_aurora import importiere_arbeitsmappe, lies_arbeitsmappe

        mappe = lies_arbeitsmappe(
            _mappe_bauen(mit_jahresabregelung=False), "Aurora.xlsx"
        )
        ergebnis = importiere_arbeitsmappe(mappe, "A", "Pult", ["Central"])[0]
        assert ergebnis.szenario.erzeugungsmenge_negativ_6h_pct_je_monat[2030] == (
            pytest.approx([0.10] * 12)
        )
        assert any("1h" in h or "6h" in h for h in ergebnis.hinweise)

    def test_mappe_ohne_monatsblatt_wird_abgewiesen(self):
        """Dieselbe harte Bedingung wie beim CSV-Weg."""
        import io as _io

        import openpyxl

        from engine.io_aurora import lies_arbeitsmappe

        wb = openpyxl.Workbook()
        wb.active.title = "Central"
        puffer = _io.BytesIO()
        wb.save(puffer)
        with pytest.raises(AuroraImportFehler, match="Monatsmodell"):
            lies_arbeitsmappe(puffer.getvalue(), "kaputt.xlsx")

    def test_unbekannte_technologie_meldet_sich(self):
        from engine.io_aurora import importiere_arbeitsmappe, lies_arbeitsmappe

        mappe = lies_arbeitsmappe(_mappe_bauen(), "Aurora.xlsx")
        with pytest.raises(AuroraImportFehler, match="Pult"):
            importiere_arbeitsmappe(mappe, "A", "Fassade", ["Central"])
